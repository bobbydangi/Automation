import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import numpy as np

random.seed(42)

# ─── Colour palette ───────────────────────────────────────────────────────────
DARK_RED   = "8B0000"
MED_RED    = "C0392B"
GOLD       = "C8922A"
WHITE      = "FFFFFF"
ALT_ROW    = "FFF5F5"
GREEN_BG   = "E8F5E9"
YELLOW_BG  = "FFFDE7"
RED_BG     = "FFEBEE"
ORANGE_BG  = "FFF3E0"
BLUE_HDR   = "1A237E"
LIGHT_BLUE = "E8EAF6"
GREY       = "F5F5F5"
SEV_CRIT   = "B71C1C"   # Critical
SEV_HIGH   = "E53935"   # High
SEV_MED    = "FB8C00"   # Medium
SEV_LOW    = "43A047"   # Low

def thin(color="CCCCCC"):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, r, c, val, bg=DARK_RED, fg=WHITE, sz=10, bold=True, wrap=True, center=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", size=sz, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=wrap)
    cell.border = thin()
    return cell

def dat(ws, r, c, val, bg=WHITE, bold=False, wrap=True, color="222222", sz=9, align="left"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", size=sz, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = thin()
    return cell

# ─── SLA THRESHOLDS per sub-category ─────────────────────────────────────────
SLA_MAP = {
    "DAM Agent Update":       3,
    "DAM Policy Change":      5,
    "DAM Access Review":      7,
    "DAM Alert Tuning":       4,
    "DAM Reporting":          7,
    "DAM Certificate Renewal":2,
    "DAM DB Onboarding":      10,
    "DAM Rule Update":        3,
}

APPROVER_POOL = [
    "Rahul Sharma", "Priya Mehta", "Anil Kumar",
    "Sunita Rao",   "Deepak Nair", "Kavita Joshi",
    "Amit Patel",   "Neha Gupta",  "Ravi Iyer",
]

DESCRIPTIONS = {
    "DAM Agent Update":
        ["Upgrade DAM agent from v3.2 to v3.5 on {ci} to address CVE-2024-1234 vulnerability.",
         "Quarterly DAM agent patch deployment on {ci} as per security baseline.",
         "Emergency agent update on {ci} post-vendor advisory for privilege escalation fix."],
    "DAM Policy Change":
        ["Modify audit policy on {ci} to include SELECT queries for PII tables per RBI directive.",
         "Update DAM policy on {ci} to exclude system batch jobs from alert threshold.",
         "Policy reconfiguration on {ci} to align with new IS audit scope Q1 2025."],
    "DAM Access Review":
        ["Periodic access review for DAM console users on {ci} – remove stale accounts.",
         "Revoke DAM admin access for 3 departed employees on {ci}.",
         "Quarterly privileged access review on {ci} as per IS Audit finding AF-2024-07."],
    "DAM Alert Tuning":
        ["Tune false-positive alerts on {ci} for after-hours batch queries.",
         "Adjust alert threshold on {ci} – reduce noise from ETL processes.",
         "Recalibrate anomaly detection rules on {ci} post-DB migration."],
    "DAM Reporting":
        ["Configure monthly DAM compliance report on {ci} for RBI submission.",
         "Fix broken report schedule on {ci} – reports not generating since 12-Jan-2025.",
         "Add new executive summary section to DAM audit report on {ci}."],
    "DAM Certificate Renewal":
        ["SSL certificate renewal for DAM server connected to {ci} – expiry in 5 days.",
         "Renew DAM collector certificate on {ci} before expiry on 28-Feb-2025.",
         "Emergency certificate replacement on {ci} after revocation by internal CA."],
    "DAM DB Onboarding":
        ["Onboard new Oracle DB {ci} to DAM monitoring scope per IS Audit recommendation.",
         "Register MySQL cluster {ci} with DAM for real-time query monitoring.",
         "Add PostgreSQL DB {ci} to DAM policy group – new application go-live."],
    "DAM Rule Update":
        ["Update detection rule on {ci} to flag bulk DELETE operations >500 rows.",
         "Add new rule on {ci} to detect login from non-whitelisted IPs.",
         "Modify masking rule on {ci} for Aadhaar number fields per UIDAI guidelines."],
}

COMMANDS = {
    "DAM Agent Update":
        "sudo systemctl stop dam-agent\nyum update dam-agent-3.5.0\nsudo systemctl start dam-agent\ndam-agent --verify",
    "DAM Policy Change":
        "dam-cli policy edit --ci {ci} --add-scope SELECT,INSERT\ndam-cli policy validate\ndam-cli policy deploy --ci {ci}",
    "DAM Alert Tuning":
        "dam-cli alerts list --ci {ci}\ndam-cli alerts set-threshold --rule BATCH_JOB --level INFO\ndam-cli alerts test --ci {ci}",
    "DAM Access Review":
        "dam-cli users list --ci {ci}\ndam-cli users revoke --user stale_account@icicibank.com\ndam-cli audit-log export --last 90d",
    "DAM Certificate Renewal":
        "openssl req -new -key dam.key -out dam.csr\nopenssl x509 -req -days 365 -in dam.csr -signkey dam.key -out dam.crt\ndam-cli cert install --cert dam.crt",
    "DAM Reporting":
        "dam-cli report schedule --ci {ci} --freq monthly --output /reports/dam/\ndam-cli report test-run --ci {ci}",
    "DAM DB Onboarding":
        "dam-cli db register --host {ci} --type oracle --port 1521\ndam-cli db test-connection --ci {ci}\ndam-cli policy assign --ci {ci} --policy DEFAULT_AUDIT",
    "DAM Rule Update":
        "dam-cli rules list --ci {ci}\ndam-cli rules update --rule BULK_DELETE --threshold 500\ndam-cli rules deploy --ci {ci}\ndam-cli rules validate",
}

CI_LIST = [
    "ICICI-ORA-PROD-001", "ICICI-ORA-PROD-002", "ICICI-MYSQL-001",
    "ICICI-MSSQL-CORE-01", "ICICI-PG-RPT-001", "ICICI-ORA-UAT-003",
    "ICICI-MYSQL-LOANS-01", "ICICI-MSSQL-CRM-02", "ICICI-ORA-CORE-005",
    "ICICI-PG-ANALYTICS-01",
]

REASONS = [
    "Mandatory security patching per IS Audit recommendation #{ref}.",
    "RBI IT Framework compliance – control {ref}.",
    "Vendor advisory {ref} – critical vulnerability remediation.",
    "Quarterly change management cycle Q{q} 2025.",
    "IS Audit finding AF-2025-{ref} – immediate remediation required.",
    "Business continuity requirement – new DB in production scope.",
    "Internal audit recommendation from {ref} review.",
]

# ─── GENERATE DUMMY DATA ──────────────────────────────────────────────────────
def generate_dummy_data(n=40):
    records = []
    base = datetime(2025, 1, 5)
    subcats = list(SLA_MAP.keys())

    for i in range(n):
        subcat   = random.choice(subcats)
        ci       = random.choice(CI_LIST)
        sla_days = SLA_MAP[subcat]

        planned_start = base + timedelta(days=random.randint(0, 110),
                                         hours=random.randint(18, 23),
                                         minutes=random.choice([0, 15, 30]))

        # ~30% chance TAT breach
        if random.random() < 0.30:
            actual_days = sla_days + random.randint(1, 10)
        else:
            actual_days = random.randint(0, sla_days)

        planned_end = planned_start + timedelta(days=actual_days,
                                                hours=random.randint(0, 4))

        risk = random.choices(
            ["High", "Medium", "Low"],
            weights=[0.25, 0.45, 0.30]
        )[0]

        # Approvals — some changes missing approval (outlier)
        n_approvers = random.choices([0, 1, 2, 3], weights=[0.08, 0.15, 0.45, 0.32])[0]
        approvers   = random.sample(APPROVER_POOL, n_approvers)
        approval_str = "; ".join([f"{a} (Approved)" for a in approvers]) if approvers else "MISSING"

        desc_tmpl = random.choice(DESCRIPTIONS[subcat])
        desc      = desc_tmpl.format(ci=ci)

        reason_tmpl = random.choice(REASONS)
        reason = reason_tmpl.format(
            ref=f"ICB-{random.randint(100,999)}",
            q=random.randint(1, 4)
        )

        cmd = COMMANDS.get(subcat, "N/A").replace("{ci}", ci)

        tat = actual_days

        records.append({
            "Change Number":     f"CHG{2025000 + i + 1}",
            "Planned Start":     planned_start,
            "Planned End":       planned_end,
            "TAT (Days)":        tat,
            "SLA (Days)":        sla_days,
            "TAT Breached":      "YES" if tat > sla_days else "NO",
            "Item / CI":         ci,
            "Sub-Category":      subcat,
            "State":             random.choices(
                                    ["Closed Complete","In Progress","Scheduled","Closed Incomplete"],
                                    weights=[0.60, 0.15, 0.15, 0.10]
                                 )[0],
            "Assigned To":       random.choice(APPROVER_POOL),
            "Risk":              risk,
            "Description":       desc,
            "Reason for Change": reason,
            "Approvals":         approval_str,
            "Commands Used":     cmd,
        })

    return pd.DataFrame(records)

# ─── SEVERITY ENGINE ──────────────────────────────────────────────────────────
def assign_severity(row):
    score = 0

    # Risk weight
    score += {"High": 4, "Medium": 2, "Low": 1}.get(row["Risk"], 0)

    # TAT breach
    if row["TAT Breached"] == "YES":
        breach = row["TAT (Days)"] - row["SLA (Days)"]
        score += 3 if breach > 5 else 2

    # Missing approvals
    if row["Approvals"] == "MISSING":
        score += 4

    # Sub-category criticality
    crit_subs = {"DAM Certificate Renewal", "DAM Agent Update", "DAM Rule Update"}
    if row["Sub-Category"] in crit_subs:
        score += 1

    # Closed incomplete
    if row["State"] == "Closed Incomplete":
        score += 2

    if score >= 9:    return "CRITICAL"
    elif score >= 6:  return "HIGH"
    elif score >= 3:  return "MEDIUM"
    else:             return "LOW"

def is_outlier(row):
    flags = []
    if row["Approvals"] == "MISSING":
        flags.append("Missing approval")
    if row["TAT Breached"] == "YES":
        flags.append(f"TAT exceeded SLA by {row['TAT (Days)'] - row['SLA (Days)']}d")
    if row["State"] == "Closed Incomplete":
        flags.append("Closed incomplete")
    if row["Risk"] == "High" and row["Approvals"] == "MISSING":
        flags.append("High risk + no approval")
    if row["TAT (Days)"] > 3 * row["SLA (Days)"]:
        flags.append("TAT > 3x SLA")
    return "; ".join(flags) if flags else ""

# ─── WRITE SHEET HELPER ───────────────────────────────────────────────────────
SEV_COLORS = {
    "CRITICAL": ("B71C1C", "FFFFFF"),
    "HIGH":     ("E53935", "FFFFFF"),
    "MEDIUM":   ("FB8C00", "FFFFFF"),
    "LOW":      ("388E3C", "FFFFFF"),
}
TAT_COLORS = {"YES": "FFEBEE", "NO": "E8F5E9"}
STATE_COLORS = {
    "Closed Complete":   "E8F5E9",
    "In Progress":       "E3F2FD",
    "Scheduled":         "FFF8E1",
    "Closed Incomplete": "FFEBEE",
}

ALL_COLS = [
    ("Change Number",     16),
    ("Planned Start",     22),
    ("Planned End",       22),
    ("TAT (Days)",        11),
    ("SLA (Days)",        11),
    ("TAT Breached",      13),
    ("Item / CI",         25),
    ("Sub-Category",      22),
    ("State",             20),
    ("Assigned To",       20),
    ("Risk",              11),
    ("Severity",          12),
    ("Outlier Flags",     40),
    ("Description",       50),
    ("Reason for Change", 45),
    ("Approvals",         40),
    ("Commands Used",     45),
]

def write_df_to_sheet(ws, df, title, title_bg=DARK_RED, show_outlier_col=True):
    ws.sheet_view.showGridLines = False

    # Title banner
    n_cols = len(ALL_COLS)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=title_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Column widths + headers
    for ci, (col_name, width) in enumerate(ALL_COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        hdr(ws, 2, ci, col_name, bg=title_bg, sz=9)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    # Data rows
    col_names = [c for c, _ in ALL_COLS]
    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        bg_base = ALT_ROW if ri % 2 == 0 else WHITE
        for ci, col_name in enumerate(col_names, start=1):
            val = row.get(col_name, "")
            if isinstance(val, pd.Timestamp):
                val = val.strftime("%d-%b-%Y %H:%M")

            bg = bg_base
            color = "222222"
            bold = False

            if col_name == "TAT Breached":
                bg = TAT_COLORS.get(str(val), bg_base)
                bold = (val == "YES")
                color = MED_RED if val == "YES" else "2E7D32"

            elif col_name == "Severity":
                sev_bg, sev_fg = SEV_COLORS.get(str(val), (bg_base, "222222"))
                bg, color, bold = sev_bg, sev_fg, True

            elif col_name == "State":
                bg = STATE_COLORS.get(str(val), bg_base)

            elif col_name == "Risk":
                if val == "High":   color, bold = MED_RED, True
                elif val == "Medium": color = "E65100"
                else: color = "2E7D32"

            elif col_name == "Outlier Flags" and val:
                bg, color, bold = "FFF3E0", "BF360C", True

            elif col_name == "Approvals" and val == "MISSING":
                bg, color, bold = "FFEBEE", MED_RED, True

            c = dat(ws, ri, ci, val, bg=bg, bold=bold, color=color, sz=9)
            ws.row_dimensions[ri].height = 52

    # Auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"

# ─── SUMMARY SHEET ────────────────────────────────────────────────────────────
def write_summary(ws, df, outliers_df):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 18

    # Banner
    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value = "ICICI Bank – IS Audit | DAM Change Report | Executive Summary"
    c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_RED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = f"Auto-generated  |  Total Records: {len(df)}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    c.font = Font(name="Calibri", size=9, italic=True, color="AAAAAA")
    c.fill = PatternFill("solid", fgColor="1A1A1A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── KPI block ────────────────────────────────────────────────────────────
    kpis = [
        ("Total Changes",            len(df),                          DARK_RED),
        ("Outliers Detected",        len(outliers_df),                 "B71C1C"),
        ("TAT Breached",             (df["TAT Breached"]=="YES").sum(),"C0392B"),
        ("Missing Approvals",        (df["Approvals"]=="MISSING").sum(),"E53935"),
        ("CRITICAL Severity",        (df["Severity"]=="CRITICAL").sum(),"B71C1C"),
        ("HIGH Severity",            (df["Severity"]=="HIGH").sum(),   "E53935"),
        ("MEDIUM Severity",          (df["Severity"]=="MEDIUM").sum(), "FB8C00"),
        ("LOW Severity",             (df["Severity"]=="LOW").sum(),    "388E3C"),
        ("Avg TAT (days)",           round(df["TAT (Days)"].mean(), 1),"1A237E"),
        ("High Risk Changes",        (df["Risk"]=="High").sum(),       "B71C1C"),
    ]
    row = 4
    hdr(ws, row, 2, "KPI", bg=DARK_RED, sz=10)
    hdr(ws, row, 3, "Value", bg=DARK_RED, sz=10)
    ws.row_dimensions[row].height = 22
    row += 1

    for label, val, color in kpis:
        c_l = ws.cell(row=row, column=2, value=label)
        c_l.font = Font(name="Calibri", size=10, bold=True, color="222222")
        c_l.fill = PatternFill("solid", fgColor=GREY)
        c_l.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_l.border = thin()

        c_v = ws.cell(row=row, column=3, value=val)
        c_v.font = Font(name="Calibri", size=13, bold=True, color=color)
        c_v.fill = PatternFill("solid", fgColor=WHITE)
        c_v.alignment = Alignment(horizontal="center", vertical="center")
        c_v.border = thin()
        ws.row_dimensions[row].height = 26
        row += 1

    # ── Sub-category breakdown ────────────────────────────────────────────────
    row += 1
    hdr(ws, row, 2, "Sub-Category", bg=BLUE_HDR, sz=9)
    hdr(ws, row, 3, "Count", bg=BLUE_HDR, sz=9)
    hdr(ws, row, 5, "TAT Breach Count", bg=BLUE_HDR, sz=9)
    hdr(ws, row, 6, "Avg TAT (days)", bg=BLUE_HDR, sz=9)
    ws.row_dimensions[row].height = 22
    row += 1

    grp = df.groupby("Sub-Category").agg(
        Count=("Change Number","count"),
        Breaches=("TAT Breached", lambda x: (x=="YES").sum()),
        AvgTAT=("TAT (Days)", "mean")
    ).reset_index()

    for i, r in grp.iterrows():
        bg = ALT_ROW if i % 2 == 0 else WHITE
        dat(ws, row, 2, r["Sub-Category"], bg=bg, bold=True)
        dat(ws, row, 3, int(r["Count"]),   bg=bg, align="center")
        bc = "FFEBEE" if r["Breaches"] > 0 else GREEN_BG
        dat(ws, row, 5, int(r["Breaches"]),   bg=bc, bold=(r["Breaches"]>0), align="center",
            color=MED_RED if r["Breaches"]>0 else "2E7D32")
        dat(ws, row, 6, round(r["AvgTAT"],1), bg=bg, align="center")
        ws.row_dimensions[row].height = 20
        row += 1

    # ── Severity breakdown ────────────────────────────────────────────────────
    row += 1
    hdr(ws, row, 2, "Severity Level", bg=BLUE_HDR, sz=9)
    hdr(ws, row, 3, "Count", bg=BLUE_HDR, sz=9)
    ws.row_dimensions[row].height = 22
    row += 1

    for sev in ["CRITICAL","HIGH","MEDIUM","LOW"]:
        cnt = (df["Severity"]==sev).sum()
        sev_bg, sev_fg = SEV_COLORS.get(sev, (WHITE, "222222"))
        c_s = ws.cell(row=row, column=2, value=sev)
        c_s.font = Font(name="Calibri", size=10, bold=True, color=sev_fg)
        c_s.fill = PatternFill("solid", fgColor=sev_bg)
        c_s.alignment = Alignment(horizontal="center", vertical="center")
        c_s.border = thin()

        c_n = ws.cell(row=row, column=3, value=cnt)
        c_n.font = Font(name="Calibri", size=12, bold=True, color=sev_fg)
        c_n.fill = PatternFill("solid", fgColor=sev_bg)
        c_n.alignment = Alignment(horizontal="center", vertical="center")
        c_n.border = thin()
        ws.row_dimensions[row].height = 24
        row += 1

# ─── MAIN ─────────────────────────────────────────────────────────────────────
print("Generating dummy data...")
df = generate_dummy_data(40)
df["Severity"]     = df.apply(assign_severity, axis=1)
df["Outlier Flags"] = df.apply(is_outlier, axis=1)

# Filtered subsets
outliers_df  = df[df["Outlier Flags"] != ""].copy()
critical_df  = df[df["Severity"] == "CRITICAL"].copy()
high_df      = df[df["Severity"] == "HIGH"].copy()
tat_df       = df[df["TAT Breached"] == "YES"].copy()
no_appr_df   = df[df["Approvals"] == "MISSING"].copy()

print(f"Total records: {len(df)}")
print(f"Outliers: {len(outliers_df)}")
print(f"CRITICAL: {len(critical_df)}  HIGH: {len(high_df)}")
print(f"TAT breached: {len(tat_df)}  Missing approval: {len(no_appr_df)}")

wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# Sheet 1 — Summary
ws_sum = wb.create_sheet("📈 Summary")
write_summary(ws_sum, df, outliers_df)

# Sheet 2 — All Changes
ws_all = wb.create_sheet("📋 All Changes")
write_df_to_sheet(ws_all, df,
    title="DAM Change Report — All Records | ICICI Bank IS Audit",
    title_bg=DARK_RED)

# Sheet 3 — Outliers
ws_out = wb.create_sheet("🚨 Outliers")
write_df_to_sheet(ws_out, outliers_df,
    title="Outlier Changes — Anomalies Requiring IS Audit Attention",
    title_bg="4A0000")

# Sheet 4 — Critical + High
crit_high = pd.concat([critical_df, high_df]).drop_duplicates()
ws_crit = wb.create_sheet("🔴 Critical & High")
write_df_to_sheet(ws_crit, crit_high,
    title="Critical & High Severity DAM Changes",
    title_bg=SEV_CRIT)

# Sheet 5 — TAT Breached
ws_tat = wb.create_sheet("⏱ TAT Breached")
write_df_to_sheet(ws_tat, tat_df,
    title="TAT Breached Changes — SLA Violations",
    title_bg="B34700")

# Sheet 6 — Missing Approvals
ws_apr = wb.create_sheet("⚠️ Missing Approvals")
write_df_to_sheet(ws_apr, no_appr_df,
    title="Changes with Missing Approvals — Governance Gap",
    title_bg="880000")

# Sheet 7 — Raw Dummy Data (source)
ws_raw = wb.create_sheet("📂 Raw Data")
ws_raw.sheet_view.showGridLines = True
raw_cols = list(df.columns)
for ci, col in enumerate(raw_cols, start=1):
    c = ws_raw.cell(row=1, column=ci, value=col)
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor="333333")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin()
    ws_raw.column_dimensions[get_column_letter(ci)].width = 20

for ri, (_, row) in enumerate(df.iterrows(), start=2):
    for ci, col in enumerate(raw_cols, start=1):
        val = row[col]
        if isinstance(val, pd.Timestamp):
            val = val.strftime("%d-%b-%Y %H:%M")
        c = ws_raw.cell(row=ri, column=ci, value=val)
        c.font = Font(name="Calibri", size=8)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = thin("DDDDDD")
        ws_raw.row_dimensions[ri].height = 30

ws_raw.freeze_panes = "A2"
ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(raw_cols))}1"

OUT = r"c:\Users\hp\Downloads\files (1)\DAM_Audit_Report.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
